# -*- coding: utf-8 -*-
"""Parcel_Comparison_Analysis



**Collection of different methods to compare tables and print out tables displaying the table changes**
"""

import pandas as pd

file1 = pd.read_csv('/content/Q2_parcel_list.csv')
file2 = pd.read_csv('/content/Q2_fp_nullOR0.csv')

# For this method you need to have the same column format. In order and amount.
# It is combining the dataframes and then removing duplicate rows.

differences = pd.concat([file1, file2]).drop_duplicates(keep=False)

print(differences)

# This method we're merging the table instead of concatenating.
# Concatenate when appending rows but same columns.
# Merge when the column numbers are different.

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# Read the tables into pandas DataFrames
table1 = pd.read_csv('/content/Q2_fp_nullOR0.csv')
table2 = pd.read_csv('/content/Q2_parcel_list.csv')

# Specify the column to ignore
column_to_ignore = 'building_fp'

# Drop the column from both tables
table1_without_column = table1.drop(columns=[column_to_ignore])

# Merge the modified tables based on common columns
merged = pd.merge(table1_without_column, table2, how='outer', indicator=True)

# Filter for differences
differences = merged[merged['_merge'] != 'both']

table1_without_column.info()

table2.info()

"""<b/> Now we're going to use the isin() function to check whether each element in the dataframe is contained in another dataframe.

And with those results, create an excel file where the variables assigned from the results of the isin function are highlighted on the excel sheet along with the updated list.</b>
"""

import pandas as pd

# Read the first CSV file
df1 = pd.read_csv('/content/Updated_Q2_NObuildings.csv')

# Read the second CSV file
df2 = pd.read_csv('/content/Q2_parcel_list.csv')

# Select only the "address" column from each DataFrame
bbl1 = df1['bbl']
bbl2 = df2['bbl']

# Find the common addresses between the two files
common_bbl = bbl1[bbl1.isin(bbl2)]

# Print the common addresses
print(common_bbl)

# Create an Excel writer object
excel_writer = pd.ExcelWriter('output13.xlsx', engine='openpyxl')

# Write the DataFrame to the Excel file
df1.to_excel(excel_writer, index=False, sheet_name='Sheet1')

# Access the workbook and sheet
workbook = excel_writer.book
worksheet = workbook['Sheet1']

# Define the fill pattern for highlighting
fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Iterate over the rows in the sheet
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
    row_values = [cell.value for cell in row]
    if row_values[4] in common_bbl.values:
        for cell in row:
            cell.fill = fill

# Save the Excel file
excel_writer.save()